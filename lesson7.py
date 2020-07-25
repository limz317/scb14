import openpyxl
import requests
# 读取测试用例函数
def read_data(filename,sheetname):#filename文件名字,sheetname表单名字，可变的，形参
    wb = openpyxl.load_workbook(filename) # 加载工作簿——文档名字，用wb这个变量接收一下工作簿workbook返回值
    sheet = wb[sheetname]# 获取表单
    max_row = sheet.max_row# 获取表单的最大行数
    case_list = [] # 创建空列表，把每次读取的字典加进去
    for i in range(2,max_row+1):#用for 循环遍历每一行
        dict1 = dict(
        case_id = sheet.cell(row=i,column=1).value, # 获取测试用例里的id
        url = sheet.cell(row=i,column=5).value, # 获取测试用例里的url
        data = sheet.cell(row=i,column=6).value, # 获取测试用例里的参数data
        expect = sheet.cell(row=i,column=7).value, # 获取测试用例里的期望
        )
        case_list.append(dict1) # 每循环一次，就把取到的每一行字典数据放到这个列表里
    return case_list # 返回测试用例列表

# 执行接口函数
def aip_fun(url,data):
    headers_reg = {'X-Lemonban-Media-Type':'lemonban.v2', 'Content-Type':'application/json'}                              #请求头
    res = requests.post(url=url,json=data,headers=headers_reg)
    response = res.json()
    return response

# 写入结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename) # 打开表单
    sheet = wb[sheetname]# 获取表单
    case_id = sheet.cell(row=row,column=column).value = final_result # 写入结果
    wb.save(filename)#保存。保存之前关闭原文档，再运行

# 封装函数
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)# 调用读取测试用例，获取所有的测试用例数据保存到变量
    for case in cases:
        case_id = case.get('case_id') #case['case_id']
        url = case.get('url')
        data = eval(case.get('data')) # eval运行被字符串包裹的表达式————去掉字符串引号
        expect = eval(case.get('expect')) # 获取预期结果，
        expect_msg = expect['msg']# 获取预期结果的msg

        real_result = aip_fun(url=url,data=data) #调用发送接口请求函数，返回结果用变量real_result接收
        real_msg = real_result['msg'] #实际结构中的msg
        print('预期结果的msg：{}'.format(expect_msg))
        print('实际结果的msg：{}'.format(real_msg))

        if expect_msg == real_msg:
            print('这{}条测试用例执行通过'.format(case_id))
            final_re = "Passed"
        else:
            print('这{}条测试用例执行不通过'.format(case_id))
            final_re = "No"
        write_result(filename,sheetname,case_id+1,8,final_re)
        print()

execute_fun("test_case_api.xlsx",'register')


# # 断言
cases = read_data("test_case_api.xlsx",'register')# 调用读取测试用例，获取所有的测试用例数据保存到变量
for case in cases:
    case_id = case.get('case_id') #case['case_id']
    url = case.get('url')
    data = eval(case.get('data'))# eval运行被字符串包裹的表达式————去掉字符串引号
    expect = eval(case.get('expect'))# 获取预期结果
    expect_msg = expect.get('msg')# 获取预期结果的msg

    real_result = aip_fun(url=url,data=data)#调用发送接口请求函数，返回结果用变量real_result接收
    real_msg = real_result.get('msg')# 获取实际结果msg

    print('预期结果的msg：{}'.format(expect_msg))
    print('实际结果的msg：{}'.format(real_msg))

    if expect_msg == real_msg:
        print('这{}条测试用例执行通过'.format(case_id))
        final_re = "Passed"
    else:
        print('这{}条测试用例执行不通过'.format(case_id))
        final_re = "No"

    write_result("test_case_api.xlsx",'register',case_id+1,8,final_re)
    print()




